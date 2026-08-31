from __future__ import annotations

import json
import base64
import csv
import io
import hashlib
import os
import re
import threading
import shutil
import traceback
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, File, Form, Header, HTTPException, UploadFile
from google.auth.transport import requests as google_auth_requests
from google.oauth2 import id_token
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import ValidationError
from starlette.concurrency import run_in_threadpool

load_dotenv()

from app.models import (  # noqa: E402
    CharacterRecord, CharacterRecastCreate, CharacterUpdate, DialogueCreate, DialogueRecord, InboxManifest,
    HistoryRename, JobEvent, JobRecord, MergeRetryCreate, ProductionCreate,
    ProjectCreate, ProjectRecord, ProjectRequest, ProjectUpdate, VoicePackDraftCreate,
    VoicePreviewCreate,
    VoiceSelectionCreate,
)
from app import state_store  # noqa: E402
from app import task_queue  # noqa: E402
from app.pipeline import (  # noqa: E402
    ARTIFACT_ROOT, CLOUD_LOCATION, DEMO_MODE, TEXT_MODEL, TTS_MODEL,
    USE_ADK_ORCHESTRATION, VOICE_EVENT_CATALOG, VOICE_EVENT_MAP, VOICE_LIBRARY,
    build_consistency_dashboard, build_export_presets, engine,
)

app = FastAPI(title="RoleVox", version="0.2.0")
STATIC = Path(__file__).parent / "static"
MAX_IMAGE_BYTES = 5 * 1024 * 1024
ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}
PROJECTS: dict[str, ProjectRecord] = {}
PROJECT_IMAGES: dict[tuple[str, str], dict] = {}
PROJECT_LOCK = threading.RLock()
INBOX_EVENT_JOBS: dict[str, str] = {}
PROJECT_ROOT = ARTIFACT_ROOT / "projects"
PRODUCTION_TARGETS = {"draft": 72, "production": 86, "cinematic": 92}


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok", "mode": engine.backend_name(),
            "configured": engine.is_configured(), "location": CLOUD_LOCATION,
            "text_model": TEXT_MODEL, "tts_model": TTS_MODEL,
            "gcs": bool(os.getenv("GCS_BUCKET")),
            "persistence": "firestore" if state_store.enabled() else "local",
            "orchestrator": "google-adk" if USE_ADK_ORCHESTRATION else "native",
            "worker": "cloud-tasks" if task_queue.enabled() else "background"}


@app.get("/api/voices")
def voices() -> dict:
    return {"voices": [{"name": key, "style": value} for key, value in VOICE_LIBRARY.items()]}


def _project_or_404(project_id: str) -> ProjectRecord:
    project = PROJECTS.get(project_id)
    if not project:
        project = state_store.get_project(project_id)
        if project:
            PROJECTS[project.id] = project
    if not project:
        raise HTTPException(404, "Project not found")
    return project


def _character_or_404(project: ProjectRecord, character_id: str) -> CharacterRecord:
    character = next((item for item in project.characters if item.id == character_id), None)
    if not character:
        raise HTTPException(404, "Character not found")
    return character


def _touch(project: ProjectRecord) -> None:
    project.updated_at = datetime.now(timezone.utc).isoformat()


def _save_project(project: ProjectRecord) -> None:
    folder = PROJECT_ROOT / project.id
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "project.json").write_text(project.model_dump_json(indent=2), encoding="utf-8")
    state_store.save_project(project)


def _ensure_voice_candidates(character: CharacterRecord) -> None:
    casting = character.casting
    candidates = casting.get("voice_candidates")
    if isinstance(candidates, list) and len(candidates) >= 3:
        return
    current = casting.get("voice") if casting.get("voice") in VOICE_LIBRARY else next(iter(VOICE_LIBRARY))
    identity = casting.get("voice_identity") if isinstance(casting.get("voice_identity"), dict) else {}
    voices = [current] + [voice for voice in VOICE_LIBRARY if voice != current][:2]
    casting["voice_candidates"] = [{"voice": voice, "label": f"OPTION {index + 1}",
        "qualities": identity.get("qualities", VOICE_LIBRARY[voice].title()) if index == 0 else VOICE_LIBRARY[voice].title(),
        "pitch": identity.get("pitch", casting.get("suggested_register", "Medium")),
        "texture": identity.get("texture", casting.get("voice_texture", VOICE_LIBRARY[voice])),
        "speaking_style": identity.get("speaking_style", casting.get("delivery_style", "Measured")),
        "accent": identity.get("accent", "Neutral"),
        "profile": casting.get("profile", f"original {VOICE_LIBRARY[voice]} fictional voice") if index == 0 else f"original {VOICE_LIBRARY[voice]} fictional voice",
        "rationale": "Existing voice identity" if index == 0 else "Additional synthetic audition option"}
        for index, voice in enumerate(voices)]
    casting.setdefault("selected_voice", current)


def _load_projects() -> None:
    if PROJECT_ROOT.exists():
        for path in PROJECT_ROOT.glob("*/project.json"):
            try:
                project = ProjectRecord.model_validate_json(path.read_text(encoding="utf-8"))
                PROJECTS[project.id] = project
                for character in project.characters:
                    _ensure_voice_candidates(character)
                    image_path = path.parent / "characters" / character.image_storage_name
                    if image_path.is_file():
                        PROJECT_IMAGES[(project.id, character.id)] = {
                            "data": image_path.read_bytes(), "mime_type": character.image_mime_type,
                            "filename": character.image_filename,
                        }
            except (OSError, ValueError):
                continue
    try:
        for project in state_store.load_projects():
            PROJECTS[project.id] = project
            for character in project.characters:
                _ensure_voice_candidates(character)
    except Exception:
        traceback.print_exc()


def _image_reference(project: ProjectRecord, character: CharacterRecord) -> dict | None:
    key = (project.id, character.id)
    reference = PROJECT_IMAGES.get(key)
    if reference:
        return reference
    data = state_store.load_character_image(project.id, character.image_storage_name)
    if data is None:
        return None
    reference = {
        "data": data,
        "mime_type": character.image_mime_type,
        "filename": character.image_filename,
    }
    PROJECT_IMAGES[key] = reference
    return reference


_load_projects()


@app.post("/api/projects", response_model=ProjectRecord, status_code=201)
def create_project(payload: ProjectCreate) -> ProjectRecord:
    project = ProjectRecord(id=uuid.uuid4().hex[:12], **payload.model_dump())
    with PROJECT_LOCK:
        PROJECTS[project.id] = project
        _save_project(project)
    return project


@app.get("/api/projects", response_model=list[ProjectRecord])
def list_projects() -> list[ProjectRecord]:
    return sorted(PROJECTS.values(), key=lambda item: item.updated_at, reverse=True)


@app.get("/api/projects/{project_id}", response_model=ProjectRecord)
def get_project(project_id: str) -> ProjectRecord:
    return _project_or_404(project_id)


@app.patch("/api/projects/{project_id}", response_model=ProjectRecord)
def update_project(project_id: str, payload: ProjectUpdate) -> ProjectRecord:
    project = _project_or_404(project_id)
    changes = payload.model_dump(exclude_none=True)
    if not changes:
        raise HTTPException(422, "Provide at least one project field to update.")
    cleaned = {key: value.strip() for key, value in changes.items()}
    if any(not value for value in cleaned.values()):
        raise HTTPException(422, "Project fields cannot be blank.")
    with PROJECT_LOCK:
        for key, value in cleaned.items():
            setattr(project, key, value)
        if any(key in cleaned for key in ("scene", "background")):
            for character in project.characters:
                character.casting.pop("preview_lines", None)
        _touch(project)
        _save_project(project)
    return project


@app.patch("/api/projects/{project_id}/characters/{character_id}", response_model=ProjectRecord)
def update_character(project_id: str, character_id: str,
                     payload: CharacterUpdate) -> ProjectRecord:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    name = payload.name.strip()
    brief = payload.brief.strip()
    if not name or not brief:
        raise HTTPException(422, "Character name and brief cannot be blank.")
    if any(item.id != character.id and item.name.casefold() == name.casefold()
           for item in project.characters):
        raise HTTPException(409, f"Character {name} already exists in this project.")
    with PROJECT_LOCK:
        character.name = name
        character.brief = brief
        character.casting["character"] = name
        character.casting.pop("preview_lines", None)
        _touch(project)
        _save_project(project)
    return project


@app.delete("/api/projects/{project_id}/characters/{character_id}", response_model=ProjectRecord)
def delete_character(project_id: str, character_id: str) -> ProjectRecord:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    image_root = (PROJECT_ROOT / project.id / "characters").resolve()
    image_path = (image_root / character.image_storage_name).resolve()
    if image_path.parent != image_root:
        raise HTTPException(400, "Invalid character image path")
    trash_root = (PROJECT_ROOT.parent / "project-trash" / "characters").resolve()
    destination = trash_root / (
        f"{project.id}_{character.id}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"
    )
    with PROJECT_LOCK:
        destination.mkdir(parents=True, exist_ok=False)
        (destination / "character.json").write_text(
            character.model_dump_json(indent=2), encoding="utf-8"
        )
        if image_path.is_file():
            shutil.move(str(image_path), str(destination / image_path.name))
        project.characters = [item for item in project.characters if item.id != character.id]
        for remaining in project.characters:
            for dialogue in remaining.dialogues:
                if dialogue.addressee_id == character.id:
                    dialogue.addressee_id = None
        PROJECT_IMAGES.pop((project.id, character.id), None)
        _touch(project)
        _save_project(project)
    return project


@app.delete("/api/projects/{project_id}", status_code=204)
def delete_project(project_id: str) -> Response:
    project = _project_or_404(project_id)
    source = (PROJECT_ROOT / project.id).resolve()
    root = PROJECT_ROOT.resolve()
    trash_root = (PROJECT_ROOT.parent / "project-trash").resolve()
    if source.parent != root:
        raise HTTPException(400, "Invalid project storage path")
    with PROJECT_LOCK:
        if source.exists():
            trash_root.mkdir(parents=True, exist_ok=True)
            destination = trash_root / f"{project.id}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"
            shutil.move(str(source), str(destination))
        PROJECTS.pop(project.id, None)
        for key in [key for key in PROJECT_IMAGES if key[0] == project.id]:
            PROJECT_IMAGES.pop(key, None)
        state_store.archive_project(project.id)
    return Response(status_code=204)


@app.get("/api/projects/{project_id}/characters/{character_id}/image")
def get_character_image(project_id: str, character_id: str) -> Response:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    reference = _image_reference(project, character)
    if not reference:
        raise HTTPException(404, "Character image not found")
    return Response(content=reference["data"], media_type=reference["mime_type"])


@app.post("/api/projects/{project_id}/characters", response_model=ProjectRecord, status_code=201)
async def add_character(
    project_id: str,
    name: str = Form(...),
    brief: str = Form(...),
    voice_presentation: str = Form("auto"),
    image: UploadFile = File(...),
) -> ProjectRecord:
    if not engine.is_configured():
        raise HTTPException(503, "Vertex AI is not configured.")
    project = _project_or_404(project_id)
    name = name.strip()
    brief = brief.strip()
    if not name or len(name) > 80 or not brief or len(brief) > 1_000:
        raise HTTPException(422, "Character name and brief are required and must fit the field limits.")
    if voice_presentation not in {"auto", "feminine", "masculine", "neutral"}:
        raise HTTPException(422, "Voice presentation must be auto, feminine, masculine, or neutral.")
    if len(project.characters) >= 10:
        raise HTTPException(422, "A maximum of 10 characters is supported per project.")
    if any(item.name.casefold() == name.casefold() for item in project.characters):
        raise HTTPException(409, f"Character {name} already exists in this project.")
    if image.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(415, "Character images must be PNG, JPEG, or WebP.")
    data = await image.read(MAX_IMAGE_BYTES + 1)
    await image.close()
    if not data or len(data) > MAX_IMAGE_BYTES:
        raise HTTPException(413, "Character image must be between 1 byte and 5 MB.")
    reference = {"data": data, "mime_type": image.content_type,
                 "filename": image.filename or "character-reference"}
    casting = await run_in_threadpool(
        engine.cast_character, project.title, project.scene, project.background,
        name, brief, reference, voice_presentation,
    )
    _ensure_voice_candidates(CharacterRecord(id="preview", name=name, brief=brief,
        image_filename=reference["filename"], image_mime_type=image.content_type,
        image_storage_name="preview", casting=casting, voice_presentation=voice_presentation))
    character_id = uuid.uuid4().hex[:10]
    extension = {"image/png": ".png", "image/jpeg": ".jpg", "image/webp": ".webp"}[image.content_type]
    storage_name = f"{character_id}{extension}"
    character = CharacterRecord(id=character_id, name=name, brief=brief,
                                image_filename=reference["filename"], image_mime_type=image.content_type,
                                image_storage_name=storage_name, casting=casting,
                                voice_presentation=voice_presentation)
    with PROJECT_LOCK:
        project.characters.append(character)
        PROJECT_IMAGES[(project.id, character.id)] = reference
        image_folder = PROJECT_ROOT / project.id / "characters"
        image_folder.mkdir(parents=True, exist_ok=True)
        (image_folder / storage_name).write_bytes(data)
        state_store.save_character_image(
            project.id, character.id, storage_name, data, image.content_type,
        )
        _touch(project)
        _save_project(project)
    return project


@app.post("/api/projects/{project_id}/characters/{character_id}/recast", response_model=ProjectRecord)
async def recast_character_voice(project_id: str, character_id: str,
                                 payload: CharacterRecastCreate) -> ProjectRecord:
    if not engine.is_configured():
        raise HTTPException(503, "Vertex AI is not configured.")
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    if character.voice_locked:
        raise HTTPException(409, "Unlock this Voice Identity before generating new audition voices.")
    reference = _image_reference(project, character)
    if not reference:
        raise HTTPException(404, "Character image not found")
    casting = await run_in_threadpool(
        engine.cast_character, project.title, project.scene, project.background,
        character.name, character.brief, reference, payload.voice_presentation,
    )
    with PROJECT_LOCK:
        character.voice_presentation = payload.voice_presentation
        character.casting = casting
        character.voice_locked = False
        _touch(project)
        _save_project(project)
    return project


@app.post("/api/projects/{project_id}/characters/{character_id}/lock", response_model=ProjectRecord)
def lock_character_voice(project_id: str, character_id: str) -> ProjectRecord:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    if not character.casting.get("selected_voice"):
        raise HTTPException(422, "Select one audition voice before locking this character.")
    with PROJECT_LOCK:
        character.voice_locked = True
        character.casting["voice_locked"] = True
        character.casting.setdefault("voice_identity", {})["locked"] = True
        _touch(project)
        _save_project(project)
    return project


@app.post("/api/projects/{project_id}/characters/{character_id}/select-voice", response_model=ProjectRecord)
def select_character_voice(project_id: str, character_id: str,
                           payload: VoiceSelectionCreate) -> ProjectRecord:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    if character.voice_locked:
        raise HTTPException(409, "Unlock this Voice Identity before changing the audition selection.")
    candidate = next((item for item in character.casting.get("voice_candidates", [])
                      if item.get("voice") == payload.voice), None)
    if not candidate:
        raise HTTPException(422, "Select one of this character's three audition voices.")
    with PROJECT_LOCK:
        character.casting["selected_voice"] = candidate["voice"]
        character.casting["voice"] = candidate["voice"]
        character.casting["profile"] = candidate.get("profile", character.casting.get("profile", ""))
        character.casting["voice_identity"] = {key: candidate.get(key, "") for key in
            ("voice", "qualities", "pitch", "texture", "speaking_style", "accent")}
        character.casting["voice_identity"]["locked"] = False
        _touch(project)
        _save_project(project)
    return project


@app.post("/api/projects/{project_id}/characters/{character_id}/voice-preview")
def preview_character_voice(project_id: str, character_id: str,
                            payload: VoicePreviewCreate) -> Response:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    candidate = next((item for item in character.casting.get("voice_candidates", [])
                      if item.get("voice") == payload.voice), None)
    if not candidate:
        raise HTTPException(422, "Unknown audition voice.")
    try:
        preview_lines = character.casting.setdefault("preview_lines", {})
        sample = preview_lines.get(payload.language)
        if not isinstance(sample, dict) or not sample.get("text"):
            sample = engine.generate_preview_line(
                project.title, project.scene, project.background,
                character.name, character.brief, payload.language,
            )
            with PROJECT_LOCK:
                character.casting.setdefault("preview_lines", {})[payload.language] = sample
                _touch(project)
                _save_project(project)
        wav = engine.preview_voice(character.name, candidate, payload.language, sample)
    except Exception as exc:
        raise HTTPException(502, f"Voice audition could not be generated: {exc}") from exc
    preview_text = base64.b64encode(sample["text"].encode("utf-8")).decode("ascii")
    return Response(content=wav, media_type="audio/wav",
                    headers={"Content-Disposition": f'inline; filename="{character.id}_{payload.voice}.wav"',
                             "X-RoleVox-Preview-Text-B64": preview_text})


@app.get("/api/voice-events")
def list_voice_events() -> list[dict]:
    return VOICE_EVENT_CATALOG


@app.post("/api/projects/{project_id}/voice-pack/draft")
def create_voice_pack_draft(project_id: str, payload: VoicePackDraftCreate) -> dict:
    if not engine.is_configured():
        raise HTTPException(503, "Vertex AI is not configured.")
    project = _project_or_404(project_id)
    character = _character_or_404(project, payload.character_id)
    selections = [item.model_dump() for item in payload.events]
    try:
        lines = engine.generate_voice_pack_draft(
            project.title, project.scene, project.background,
            character.name, character.brief, payload.language, selections,
        )
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(502, f"Voice Pack Writer could not create the draft: {exc}") from exc
    return {"character_id": character.id, "character": character.name,
            "language": payload.language, "lines": lines}


@app.post("/api/projects/{project_id}/characters/{character_id}/unlock", response_model=ProjectRecord)
def unlock_character_voice(project_id: str, character_id: str) -> ProjectRecord:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    with PROJECT_LOCK:
        character.voice_locked = False
        character.casting["voice_locked"] = False
        character.casting.setdefault("voice_identity", {})["locked"] = False
        _touch(project)
        _save_project(project)
    return project


@app.post("/api/projects/{project_id}/characters/{character_id}/dialogues",
          response_model=ProjectRecord, status_code=201)
def add_dialogue(project_id: str, character_id: str, payload: DialogueCreate) -> ProjectRecord:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    if payload.addressee_id:
        _character_or_404(project, payload.addressee_id)
        if payload.addressee_id == character.id:
            raise HTTPException(422, "Choose another character as addressee, or let AI infer it.")
    with PROJECT_LOCK:
        next_order = max((line.order for item in project.characters for line in item.dialogues), default=0) + 1
        character.dialogues.append(DialogueRecord(id=uuid.uuid4().hex[:10], order=next_order,
                                                   **payload.model_dump()))
        _touch(project)
        _save_project(project)
    return project


def _normalize_import_rows(filename: str, data: bytes) -> list[dict[str, str]]:
    suffix = Path(filename).suffix.casefold()
    rows: list[dict] = []
    if suffix == ".json":
        payload = json.loads(data.decode("utf-8-sig"))
        rows = payload.get("lines", []) if isinstance(payload, dict) else payload
        if not isinstance(rows, list):
            raise ValueError("JSON must be an array or an object with a lines array.")
    elif suffix == ".csv":
        rows = list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))
    elif suffix == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip().casefold() for value in values[0]]
        rows = [dict(zip(headers, row)) for row in values[1:]]
    elif suffix in {".ink", ".yarn", ".txt"}:
        for raw in data.decode("utf-8-sig").splitlines():
            line = raw.strip()
            if not line or line.startswith(("//", "#", "===", "---", "<<", "title:", "tags:")):
                continue
            parts = re.split(r"[:：]", line, maxsplit=1)
            if len(parts) == 2 and parts[0].strip() and parts[1].strip():
                rows.append({"character": parts[0].strip(), "text": parts[1].strip()})
    else:
        raise ValueError("Supported formats: CSV, JSON, XLSX, Ink, Yarn, and TXT.")

    normalized = []
    for index, raw in enumerate(rows, start=2):
        if not isinstance(raw, dict):
            raise ValueError(f"Row {index} must be an object.")
        lowered = {str(key).strip().casefold(): str(value or "").strip()
                   for key, value in raw.items()}
        character = lowered.get("character") or lowered.get("speaker") or lowered.get("角色")
        text = lowered.get("text") or lowered.get("dialogue") or lowered.get("台詞") or lowered.get("文本")
        emotion = lowered.get("emotion") or lowered.get("voice emotion") or lowered.get("情緒") or "Context-aware · natural · character-authentic"
        addressee = lowered.get("addressee") or lowered.get("speaking_to") or lowered.get("speaking to") or lowered.get("對象") or ""
        if not character or not text:
            raise ValueError(f"Row {index} requires character/speaker and text/dialogue.")
        normalized.append({"character": character, "text": text, "emotion": emotion,
                           "addressee": addressee})
    if len(normalized) > 500:
        raise ValueError("Import is limited to 500 dialogue lines per file.")
    return normalized


@app.post("/api/projects/{project_id}/dialogues/import")
async def import_project_dialogues(project_id: str, file: UploadFile = File(...)) -> ProjectRecord:
    project = _project_or_404(project_id)
    data = await file.read()
    if not data or len(data) > 5 * 1024 * 1024:
        raise HTTPException(422, "Script file must be between 1 byte and 5 MB.")
    try:
        rows = _normalize_import_rows(file.filename or "script.txt", data)
    except (ValueError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise HTTPException(422, f"Script import failed: {exc}") from exc
    if not rows:
        raise HTTPException(422, "No dialogue lines were found in the imported file.")
    by_name = {item.name.casefold(): item for item in project.characters}
    missing = sorted({row["character"] for row in rows if row["character"].casefold() not in by_name})
    if missing:
        raise HTTPException(422, f"Create these Character Cards before import: {', '.join(missing)}")
    with PROJECT_LOCK:
        next_order = max((line.order for item in project.characters for line in item.dialogues), default=0) + 1
        for row in rows:
            character = by_name[row["character"].casefold()]
            addressee = by_name.get(row["addressee"].casefold()) if row["addressee"] else None
            character.dialogues.append(DialogueRecord(
                id=uuid.uuid4().hex[:10], order=next_order, emotion=row["emotion"][:80],
                text=row["text"][:4000], addressee_id=addressee.id if addressee else None,
            ))
            next_order += 1
        _touch(project)
        _save_project(project)
    return project


@app.patch("/api/projects/{project_id}/characters/{character_id}/dialogues/{dialogue_id}",
           response_model=ProjectRecord)
def update_dialogue(project_id: str, character_id: str, dialogue_id: str,
                    payload: DialogueCreate) -> ProjectRecord:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    dialogue = next((item for item in character.dialogues if item.id == dialogue_id), None)
    if not dialogue:
        raise HTTPException(404, "Dialogue not found")
    if payload.addressee_id:
        _character_or_404(project, payload.addressee_id)
        if payload.addressee_id == character.id:
            raise HTTPException(422, "Choose another character as addressee, or let AI infer it.")
    with PROJECT_LOCK:
        dialogue.emotion = payload.emotion
        dialogue.text = payload.text
        dialogue.addressee_id = payload.addressee_id
        _touch(project)
        _save_project(project)
    return project


@app.delete("/api/projects/{project_id}/characters/{character_id}/dialogues/{dialogue_id}",
            response_model=ProjectRecord)
def delete_dialogue(project_id: str, character_id: str, dialogue_id: str) -> ProjectRecord:
    project = _project_or_404(project_id)
    character = _character_or_404(project, character_id)
    with PROJECT_LOCK:
        before = len(character.dialogues)
        character.dialogues = [item for item in character.dialogues if item.id != dialogue_id]
        if len(character.dialogues) == before:
            raise HTTPException(404, "Dialogue not found")
        _touch(project)
        _save_project(project)
    return project


def _prepare_project_production(
    project_id: str, payload: ProductionCreate, *, job_id: str | None = None,
    run_origin: str = "studio",
) -> tuple[JobRecord, ProjectRequest, dict[str, dict]]:
    if not engine.is_configured():
        raise HTTPException(503, "Vertex AI is not configured.")
    project = _project_or_404(project_id)
    if not project.characters:
        raise HTTPException(422, "Add at least one character before production.")
    if payload.workflow_mode == "single":
        single_character_id = payload.single_character_id or payload.character_id
        selected_characters = [item for item in project.characters if item.id == single_character_id]
        if not selected_characters:
            raise HTTPException(404, "Select a character for single-character generation.")
        single_text = (payload.single_text or "").strip()
        single_emotion = (payload.single_emotion or "").strip()
        if not single_text or not single_emotion:
            raise HTTPException(422, "Single-character generation requires emotion and dialogue text.")
    elif payload.workflow_mode == "voice_pack":
        selected_characters = [item for item in project.characters
                               if item.id == payload.pack_character_id]
        if not selected_characters:
            raise HTTPException(404, "Select a character for voice-pack generation.")
        if not payload.pack_lines:
            raise HTTPException(422, "Generate and approve at least one voice-pack draft line.")
        seen_variants: set[tuple[str, int]] = set()
        for line in payload.pack_lines:
            event = VOICE_EVENT_MAP.get(line.event)
            if not event or line.event_label != event["label"]:
                raise HTTPException(422, f"Unknown voice-pack event: {line.event}")
            key = (line.event, line.variant)
            if key in seen_variants:
                raise HTTPException(422, f"Duplicate voice-pack variant: {line.event} {line.variant}")
            seen_variants.add(key)
    else:
        selected_characters = [item for item in project.characters if item.dialogues]
        if payload.character_id:  # Backward-compatible single-speaker script production.
            selected_characters = [item for item in selected_characters if item.id == payload.character_id]
        if not selected_characters:
            raise HTTPException(422, "Add at least one project dialogue line before dialogue production.")
    unlocked = [item.name for item in selected_characters if not item.voice_locked]
    if unlocked:
        raise HTTPException(422, f"Lock every voice before production: {', '.join(unlocked)}")
    script_lines: list[str] = []
    line_emotions: dict[int, str] = {}
    line_addressees: dict[int, str] = {}
    line_events: dict[int, str] = {}
    line_variants: dict[int, int] = {}
    line_id = 0
    if payload.workflow_mode == "single":
        character = selected_characters[0]
        script_lines.append(f"{character.name}: {single_text}")
        line_emotions[1] = single_emotion
    elif payload.workflow_mode == "voice_pack":
        character = selected_characters[0]
        for line_id, line in enumerate(payload.pack_lines, start=1):
            script_lines.append(f"{character.name}: {line.text}")
            line_emotions[line_id] = line.emotion
            line_events[line_id] = line.event
            line_variants[line_id] = line.variant
    else:
        selected_ids = {item.id for item in selected_characters}
        character_names = {item.id: item.name for item in project.characters}
        dialogue_rows = [(character, dialogue, stable_index)
                         for stable_index, (character, dialogue) in enumerate(
                             (pair for item in project.characters for pair in
                              ((item, dialogue) for dialogue in item.dialogues)))
                         if character.id in selected_ids]
        dialogue_rows.sort(key=lambda row: (row[1].order if row[1].order > 0 else 1_000_000 + row[2]))
        for character, dialogue, _ in dialogue_rows:
            line_id += 1
            script_lines.append(f"{character.name}: {dialogue.text}")
            line_emotions[line_id] = dialogue.emotion
            if dialogue.addressee_id and dialogue.addressee_id in character_names:
                line_addressees[line_id] = character_names[dialogue.addressee_id]
    request = ProjectRequest(
        title=project.title, scene=project.scene, background=project.background,
        target_language=payload.target_language, script="\n".join(script_lines),
        character_descriptions={item.name: item.brief for item in selected_characters},
        quality_threshold=PRODUCTION_TARGETS[payload.production_mode],
        max_retries=payload.revision_limit, production_mode=payload.production_mode,
        workflow_mode=payload.workflow_mode,
        line_emotions=line_emotions, line_addressees=line_addressees,
        line_events=line_events, line_variants=line_variants,
        locked_casting=[item.casting for item in selected_characters],
        run_origin=run_origin,
    )
    character_images = {
        item.name: reference
        for item in selected_characters
        if (reference := _image_reference(project, item)) is not None
    }
    job = engine.create(job_id or uuid.uuid4().hex[:12], request, project.id)
    return job, request, character_images


@app.post("/api/projects/{project_id}/produce", response_model=JobRecord, status_code=202)
def produce_project(project_id: str, payload: ProductionCreate,
                    tasks: BackgroundTasks) -> JobRecord:
    job, request, character_images = _prepare_project_production(project_id, payload)
    if task_queue.enabled():
        queue_name = task_queue.enqueue(job.id, request, project_id)
        job.events.append(JobEvent(
            agent="Cloud Tasks Dispatcher", message=f"Queued durable worker task: {queue_name}",
            status="info",
        ))
        state_store.save_job(job)
    else:
        tasks.add_task(engine.run, job, request, character_images)
    return job


def _verify_google_caller(authorization: str | None, expected_account: str,
                          expected_audience: str | list[str], purpose: str) -> dict:
    has_audience = any(expected_audience) if isinstance(expected_audience, list) else bool(expected_audience)
    if not expected_account or not has_audience:
        raise HTTPException(503, f"RoleVox {purpose} authentication is not configured.")
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "A Google OIDC bearer token is required.")
    try:
        claims = id_token.verify_oauth2_token(
            authorization.removeprefix("Bearer ").strip(),
            google_auth_requests.Request(),
            audience=expected_audience,
        )
    except (ValueError, TypeError) as exc:
        raise HTTPException(401, "Invalid Google OIDC bearer token.") from exc
    if claims.get("email") != expected_account or claims.get("email_verified") is not True:
        raise HTTPException(403, f"OIDC identity is not authorized for the RoleVox {purpose}.")
    return claims


def _verify_inbox_caller(authorization: str | None) -> dict:
    expected_account = os.getenv("EVENTARC_SERVICE_ACCOUNT", "").strip()
    expected_audience = os.getenv("ROLEVOX_EVENT_AUDIENCE", "").strip()
    public_url = os.getenv("ROLEVOX_PUBLIC_URL", "").strip().rstrip("/")
    audiences = [expected_audience]
    if public_url and public_url != expected_audience:
        audiences.extend([public_url, f"{public_url}/api/inbox/events"])
    audiences.append(f"{expected_audience.rstrip('/')}/api/inbox/events")
    return _verify_google_caller(authorization, expected_account, audiences, "inbox")


def _verify_task_caller(authorization: str | None) -> dict:
    expected_account = os.getenv("CLOUD_TASKS_SERVICE_ACCOUNT", "").strip()
    expected_audience = os.getenv("ROLEVOX_EVENT_AUDIENCE", "").strip()
    return _verify_google_caller(
        authorization, expected_account, expected_audience, "production worker",
    )


def _load_inbox_manifest(bucket_name: str, object_name: str,
                         generation: str | None) -> InboxManifest:
    from google.cloud import storage

    blob_generation = int(generation) if generation and generation.isdigit() else None
    raw = storage.Client().bucket(bucket_name).blob(
        object_name, generation=blob_generation,
    ).download_as_bytes()
    try:
        return InboxManifest.model_validate_json(raw)
    except ValidationError as exc:
        raise HTTPException(422, f"Invalid RoleVox inbox manifest: {exc}") from exc


@app.post("/api/inbox/events")
def receive_inbox_event(
    event: dict, authorization: str | None = Header(default=None),
) -> dict:
    """Process one authenticated Cloud Storage finalized event synchronously."""

    _verify_inbox_caller(authorization)
    data = event.get("data") if isinstance(event.get("data"), dict) else event
    bucket_name = str(data.get("bucket", "")).strip()
    object_name = str(data.get("name", "")).strip()
    generation = str(data.get("generation", "")).strip() or None
    configured_bucket = os.getenv("GCS_BUCKET", "").strip()
    if not configured_bucket:
        raise HTTPException(503, "RoleVox inbox requires GCS_BUCKET.")
    if bucket_name != configured_bucket:
        raise HTTPException(403, "Storage event bucket is not authorized for RoleVox.")
    if not object_name.startswith("inbox/") or not object_name.lower().endswith(".json"):
        return {"status": "ignored", "reason": "Object is outside the RoleVox inbox."}

    event_key = f"{bucket_name}/{object_name}#{generation or 'latest'}"
    existing_job_id = INBOX_EVENT_JOBS.get(event_key)
    existing_job = engine.get(existing_job_id) if existing_job_id else None
    if existing_job and existing_job.status == "completed":
        return {"status": "duplicate", "job_id": existing_job.id,
                "cloud_url": (existing_job.result or {}).get("cloud_url")}

    manifest = _load_inbox_manifest(bucket_name, object_name, generation)
    production = ProductionCreate(
        target_language=manifest.target_language,
        production_mode=manifest.production_mode,
        workflow_mode="dialogue",
        revision_limit=manifest.revision_limit,
        character_id=manifest.character_id,
    )
    job_id = uuid.uuid5(uuid.NAMESPACE_URL, event_key).hex[:12]
    claimed, persisted_job_id = state_store.claim_inbox_event(event_key, job_id)
    if not claimed:
        persisted_job = engine.get(persisted_job_id) if persisted_job_id else None
        if persisted_job and persisted_job.status == "completed":
            return {"status": "duplicate", "job_id": persisted_job.id,
                    "cloud_url": (persisted_job.result or {}).get("cloud_url")}
        return {"status": "processing", "job_id": persisted_job_id or job_id}
    job, request, character_images = _prepare_project_production(
        manifest.project_id, production, job_id=job_id, run_origin="eventarc-inbox",
    )
    job.events.append(JobEvent(
        agent="Auto Production Inbox",
        message=f"Accepted Cloud Storage event for gs://{bucket_name}/{object_name}",
        status="info",
    ))
    INBOX_EVENT_JOBS[event_key] = job.id
    engine.run(job, request, character_images)
    if job.status != "completed":
        INBOX_EVENT_JOBS.pop(event_key, None)
        state_store.release_inbox_event(event_key)
        raise HTTPException(503, f"Inbox production failed: {job.error or 'unknown error'}")
    state_store.complete_inbox_event(event_key, job.id)
    return {"status": "completed", "job_id": job.id,
            "cloud_url": (job.result or {}).get("cloud_url")}


@app.post("/api/jobs", response_model=JobRecord, status_code=202)
def create_job(payload: ProjectRequest, tasks: BackgroundTasks) -> JobRecord:
    if not engine.is_configured():
        raise HTTPException(503, "Configure Vertex AI/ADC or GEMINI_API_KEY, or explicitly enable DEMO_MODE=true.")
    payload.run_origin = "api"
    job = engine.create(uuid.uuid4().hex[:12], payload)
    if task_queue.enabled():
        queue_name = task_queue.enqueue(job.id, payload)
        job.events.append(JobEvent(
            agent="Cloud Tasks Dispatcher", message=f"Queued durable worker task: {queue_name}",
            status="info",
        ))
        state_store.save_job(job)
    else:
        tasks.add_task(engine.run, job, payload)
    return job


@app.post("/api/jobs/{job_id}/execute")
def execute_job(job_id: str, payload: dict,
                authorization: str | None = Header(default=None)) -> dict:
    """Authenticated Cloud Tasks target that keeps CPU allocated until production ends."""

    _verify_task_caller(authorization)
    job = engine.get(job_id)
    if not job:
        raise HTTPException(404, "Queued job not found")
    if job.status == "completed":
        return {"status": "duplicate", "job_id": job.id}
    try:
        request = ProjectRequest.model_validate(payload.get("request"))
    except ValidationError as exc:
        raise HTTPException(422, f"Invalid queued production request: {exc}") from exc
    character_images: dict[str, dict] = {}
    project_id = payload.get("project_id")
    if project_id:
        project = _project_or_404(str(project_id))
        for character in project.characters:
            if character.name in request.character_descriptions:
                reference = _image_reference(project, character)
                if reference:
                    character_images[character.name] = reference
    engine.run(job, request, character_images)
    return {"status": job.status, "job_id": job.id, "error": job.error}


def _script_characters(script: str) -> list[str]:
    names: list[str] = []
    for raw in script.splitlines():
        parts = re.split(r"[:：]", raw.strip(), maxsplit=1)
        if len(parts) == 2 and parts[0].strip() and parts[0].strip() not in names:
            names.append(parts[0].strip())
    return names


@app.post("/api/jobs/with-references", response_model=JobRecord, status_code=202)
async def create_job_with_references(
    tasks: BackgroundTasks,
    payload: str = Form(...),
    image_characters: str = Form("[]"),
    images: list[UploadFile] = File(default=[]),
) -> JobRecord:
    if not engine.is_configured():
        raise HTTPException(503, "Configure Vertex AI/ADC or GEMINI_API_KEY, or explicitly enable DEMO_MODE=true.")
    try:
        request = ProjectRequest.model_validate_json(payload)
        mapped_characters = json.loads(image_characters)
    except (ValidationError, json.JSONDecodeError) as exc:
        raise HTTPException(422, f"Invalid production payload: {exc}") from exc
    if not isinstance(mapped_characters, list) or len(mapped_characters) != len(images):
        raise HTTPException(422, "Each uploaded image must map to exactly one character.")
    script_characters = _script_characters(request.script)
    if len(script_characters) > 10:
        raise HTTPException(422, "A maximum of 10 characters is supported per production.")
    unknown = (set(request.character_descriptions) | set(mapped_characters)) - set(script_characters)
    if unknown:
        raise HTTPException(422, f"References do not match script characters: {', '.join(sorted(unknown))}")

    character_images: dict[str, dict] = {}
    for character, upload in zip(mapped_characters, images):
        if character in character_images:
            raise HTTPException(422, f"Only one image is allowed for {character}.")
        if upload.content_type not in ALLOWED_IMAGE_TYPES:
            raise HTTPException(415, "Character images must be PNG, JPEG, or WebP.")
        data = await upload.read(MAX_IMAGE_BYTES + 1)
        await upload.close()
        if not data or len(data) > MAX_IMAGE_BYTES:
            raise HTTPException(413, "Each character image must be between 1 byte and 5 MB.")
        character_images[character] = {
            "data": data,
            "mime_type": upload.content_type,
            "filename": upload.filename or "character-reference",
        }

    job = engine.create(uuid.uuid4().hex[:12], request)
    if task_queue.enabled():
        # This compatibility endpoint carries raw image bytes, so keep the request open
        # until they have been consumed rather than placing them in a task payload.
        engine.run(job, request, character_images)
    else:
        tasks.add_task(engine.run, job, request, character_images)
    return job


@app.get("/api/jobs/{job_id}", response_model=JobRecord)
def get_job(job_id: str) -> JobRecord:
    job = engine.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    return job


@app.get("/api/projects/{project_id}/jobs")
def list_project_jobs(project_id: str) -> list[dict]:
    _project_or_404(project_id)
    combined = {job.id: job for job in state_store.list_jobs(project_id)}
    combined.update({job.id: job for job in engine.jobs.values() if job.project_id == project_id})
    jobs = sorted(
        (job for job in combined.values() if not job.history_hidden),
        key=lambda item: item.created_at, reverse=True,
    )[:30]
    return [{
        "id": job.id, "status": job.status, "stage": job.stage,
        "history_name": job.history_name,
        "progress": job.progress, "workflow_mode": job.workflow_mode,
        "created_at": job.created_at, "updated_at": job.updated_at,
        "line_count": len((job.result or {}).get("lines", [])),
        "needs_review_count": int((job.result or {}).get("needs_review_count", 0)),
        "target_language": (job.result or {}).get("target_language"),
        "production_mode": (job.result or {}).get("production_mode"),
        "package_url": f"/api/jobs/{job.id}/package" if job.status == "completed" else None,
        "error": job.error,
    } for job in jobs]


@app.patch("/api/projects/{project_id}/jobs/{job_id}/history", response_model=JobRecord)
def rename_project_job_history(project_id: str, job_id: str,
                               payload: HistoryRename) -> JobRecord:
    _project_or_404(project_id)
    job = engine.get(job_id)
    if not job or job.project_id != project_id or job.history_hidden:
        raise HTTPException(404, "Production history record not found.")
    job.history_name = payload.name.strip()
    job.updated_at = datetime.now(timezone.utc).isoformat()
    state_store.save_job(job)
    return job


@app.delete("/api/projects/{project_id}/jobs/{job_id}/history", status_code=204)
def delete_project_job_history(project_id: str, job_id: str) -> Response:
    _project_or_404(project_id)
    job = engine.get(job_id)
    if not job or job.project_id != project_id:
        raise HTTPException(404, "Production history record not found.")
    if job.status in {"queued", "running"}:
        raise HTTPException(409, "A running production cannot be cleared from history.")
    job_dir = ARTIFACT_ROOT / job.id
    if job_dir.exists():
        shutil.rmtree(job_dir)
    state_store.delete_job_artifacts(job.id)
    state_store.delete_job(job.id)
    engine.jobs.pop(job.id, None)
    return Response(status_code=204)


def _safe_file(job_id: str, filename: str) -> Path:
    job_dir = (ARTIFACT_ROOT / job_id).resolve()
    path = (job_dir / filename).resolve()
    if path.parent != job_dir or not path.is_file():
        raise HTTPException(404, "File not found")
    return path


def _artifact_bytes(job_id: str, filename: str) -> bytes:
    try:
        return _safe_file(job_id, filename).read_bytes()
    except HTTPException:
        data = state_store.load_artifact(job_id, filename)
        if data is None:
            raise HTTPException(404, f"Artifact not found: {filename}")
        return data


def _persist_artifact(job_id: str, filename: str, data: bytes,
                      content_type: str = "application/octet-stream") -> None:
    job_dir = ARTIFACT_ROOT / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    (job_dir / filename).write_bytes(data)
    state_store.save_artifact(job_id, filename, data, content_type)


@app.post("/api/jobs/{job_id}/lines/{line_id}/merge-retry", response_model=JobRecord)
def merge_retry(job_id: str, line_id: int, payload: MergeRetryCreate) -> JobRecord:
    """Merge a successful one-line retry into the original durable package."""
    original = engine.get(job_id)
    replacement = engine.get(payload.replacement_job_id)
    if not original or original.status != "completed" or not original.result:
        raise HTTPException(409, "The original completed run is not available.")
    if not replacement or replacement.status != "completed" or not replacement.result:
        raise HTTPException(409, "The replacement run has not completed.")
    if original.project_id != replacement.project_id:
        raise HTTPException(409, "Retry and original run belong to different projects.")
    replacement_lines = replacement.result.get("lines", [])
    if len(replacement_lines) != 1:
        raise HTTPException(422, "A merge retry must contain exactly one line.")
    line_index = next((index for index, line in enumerate(original.result.get("lines", []))
                       if int(line.get("id", -1)) == line_id), None)
    if line_index is None:
        raise HTTPException(404, "Original line not found.")

    old = original.result["lines"][line_index]
    fresh = replacement_lines[0]
    final_name = old["file"]
    _persist_artifact(job_id, final_name, _artifact_bytes(replacement.id, fresh["file"]), "audio/wav")
    merged_takes = []
    stem = Path(final_name).stem
    for index, take in enumerate(fresh.get("takes", []), 1):
        take_name = f"{stem}_retry_take{index:02d}.wav"
        _persist_artifact(job_id, take_name, _artifact_bytes(replacement.id, take["file"]), "audio/wav")
        merged_takes.append({**take, "file": take_name,
                             "url": f"/api/jobs/{job_id}/files/{take_name}"})
    merged = {
        **old, **fresh, "id": old["id"], "file": final_name,
        "url": f"/api/jobs/{job_id}/files/{final_name}", "takes": merged_takes,
        "voice_event": old.get("voice_event"), "voice_variant": old.get("voice_variant"),
        "merged_retry_job_id": replacement.id,
    }
    original.result["lines"][line_index] = merged
    lines = original.result["lines"]
    original.result["needs_review_count"] = sum(1 for line in lines if line.get("needs_review"))
    original.result["voice_consistency"] = build_consistency_dashboard(lines)
    exports = build_export_presets(
        original.result.get("project", original.title), original.result.get("scene", ""),
        original.result.get("target_language", "en"), lines,
    )
    for name, data in exports.items():
        _persist_artifact(job_id, name, data,
                          "text/csv" if name.endswith(".csv") else "application/json")

    receipt = original.result.get("run_receipt", {})
    receipt["needs_review_count"] = original.result["needs_review_count"]
    receipt["updated_at"] = datetime.now(timezone.utc).isoformat()
    receipt["merged_retries"] = [*receipt.get("merged_retries", []),
                                 {"line_id": line_id, "replacement_job_id": replacement.id}]
    receipt_lines = []
    for line in lines:
        audio = _artifact_bytes(job_id, line["file"])
        receipt_lines.append({
            "line_id": line["id"], "character": line["character"], "voice": line["voice"],
            "approved": line["approved"], "needs_review": line["needs_review"],
            "best_available": line.get("best_available", False),
            "selected_take": line["selected_take"], "critic_score": int(line["qa"].get("score", 0)),
            "attempts": line["attempts"], "output_file": line["file"],
            "sha256": hashlib.sha256(audio).hexdigest(),
        })
    receipt["lines"] = receipt_lines
    original.result["run_receipt"] = receipt
    manifest = {key: value for key, value in original.result.items()
                if key not in {"package_url", "package_name", "cloud_url", "run_receipt"}}
    manifest["autonomous_run_receipt"] = "run_receipt.json"
    manifest_bytes = json.dumps(manifest, ensure_ascii=False, indent=2).encode()
    receipt_bytes = json.dumps(receipt, ensure_ascii=False, indent=2).encode()
    _persist_artifact(job_id, "manifest.json", manifest_bytes, "application/json")
    _persist_artifact(job_id, "run_receipt.json", receipt_bytes, "application/json")

    package_name = original.result["package_name"]
    package_path = ARTIFACT_ROOT / job_id / package_name
    with zipfile.ZipFile(package_path, "w", zipfile.ZIP_DEFLATED) as package:
        for line in lines:
            package.writestr(line["file"], _artifact_bytes(job_id, line["file"]))
        package.writestr("manifest.json", manifest_bytes)
        package.writestr("run_receipt.json", receipt_bytes)
        for name, data in exports.items():
            package.writestr(name, data)
    state_store.save_artifact(job_id, package_name, package_path.read_bytes(), "application/zip")
    original.updated_at = datetime.now(timezone.utc).isoformat()
    original.events.append(JobEvent(
        agent="Package Merge Agent",
        message=f"Line {line_id} retry merged into the original package; manifests and hashes rebuilt.",
        status="passed",
    ))
    state_store.save_job(original)
    return original


@app.get("/api/jobs/{job_id}/files/{filename}")
def get_audio(job_id: str, filename: str) -> Response:
    try:
        return FileResponse(_safe_file(job_id, filename), media_type="audio/wav", filename=filename)
    except HTTPException:
        job = engine.get(job_id)
        allowed = {
            take.get("file")
            for line in ((job.result or {}).get("lines", []) if job else [])
            for take in [line, *line.get("takes", [])]
        }
        if filename not in allowed:
            raise HTTPException(404, "File not found")
        data = state_store.load_artifact(job_id, filename)
        if data is None:
            raise HTTPException(404, "File not found")
        return Response(content=data, media_type="audio/wav",
                        headers={"Content-Disposition": f'inline; filename="{filename}"'})


@app.get("/api/jobs/{job_id}/package")
def get_package(job_id: str) -> Response:
    job = engine.get(job_id)
    if not job or job.status != "completed" or not job.result:
        raise HTTPException(404, "Completed package not found")
    filename = job.result["package_name"]
    try:
        path = _safe_file(job_id, filename)
        return FileResponse(path, media_type="application/zip", filename=path.name)
    except HTTPException:
        data = state_store.load_artifact(job_id, filename)
        if data is None:
            raise HTTPException(404, "Completed package not found")
        return Response(content=data, media_type="application/zip",
                        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/api/jobs/{job_id}/exports/{filename}")
def get_export(job_id: str, filename: str) -> Response:
    job = engine.get(job_id)
    allowed = {item.get("file") for item in ((job.result or {}).get("export_presets", []) if job else [])}
    if filename not in allowed:
        raise HTTPException(404, "Export preset not found")
    data = _artifact_bytes(job_id, filename)
    media_type = "text/csv; charset=utf-8" if filename.endswith(".csv") else "application/json"
    return Response(content=data, media_type=media_type,
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


app.mount("/", StaticFiles(directory=STATIC, html=True), name="static")
